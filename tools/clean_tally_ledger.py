"""Standalone cleaner for Tally-exported Excel ledgers.

Tally ledger exports are semi-structured: each transaction spans multiple rows
(a "To/By" line plus sub-rows for the instrument/narration) and the header row
does not column-align with the data, so amounts leak into the wrong columns.
RAG and pandas both struggle with that.

This script flattens such ledgers into a TIDY table —
  Date | Type | Account | VoucherType | Debit | Credit | Details
— and writes <name>_clean.xlsx next to the original. Clean (already-tabular)
sheets are passed through with simple header detection.

This is a one-off preprocessing tool. It is completely separate from the app,
so it cannot affect the app's quality or performance. Run it on your ledger
files, then upload the *_clean.xlsx versions.

Usage:
    python tools/clean_tally_ledger.py <file.xlsx> [more files...]
    python tools/clean_tally_ledger.py <folder>           # all .xlsx/.xls in folder
    python tools/clean_tally_ledger.py                    # defaults to local_testing/excel
"""
import sys
import os
from pathlib import Path

import pandas as pd
import openpyxl

_DRCR = {"to", "by"}
_HEADER_HINTS = {"date", "debit", "credit", "particulars", "vch", "amount"}


def _num(v):
    """Parse a cell to a number, tolerating commas; NaN if not numeric."""
    if v is None:
        return float("nan")
    return pd.to_numeric(str(v).replace(",", "").strip(), errors="coerce")


def _is_date(v) -> bool:
    if v is None:
        return False
    d = pd.to_datetime(v, errors="coerce")
    return pd.notna(d) and d.year >= 2000


def _to_date(v):
    d = pd.to_datetime(v, errors="coerce")
    return d.date().isoformat() if pd.notna(d) else None


def _largest_number(cells):
    """The transaction amount is the largest number in the row — bigger than a
    voucher number, and robust to amounts that leaked into the Date cell."""
    best = None
    for v in cells:
        n = _num(v)
        if pd.notna(n) and (best is None or n > best):
            best = float(n)
    return best


def _sheet_rows(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]


def _find_header(rows) -> int:
    for i, row in enumerate(rows[:30]):
        low = {str(c).strip().lower() for c in row if c is not None}
        if "date" in low and len(low & _HEADER_HINTS) >= 2:
            return i
    return 0


def _metadata(rows, header_i) -> str:
    parts = []
    for row in rows[:header_i]:
        for c in row:
            if c is not None and str(c).strip():
                parts.append(str(c).strip())
    return " | ".join(parts[:6])


def _is_tally(rows, header_i) -> bool:
    """True if column B is dominated by To/By markers (Tally party-ledger)."""
    tb = tot = 0
    for row in rows[header_i + 1:]:
        b = row[1] if len(row) > 1 else None
        if b is not None and str(b).strip():
            tot += 1
            if str(b).strip().lower() in _DRCR:
                tb += 1
    return tot > 0 and tb / tot > 0.3


def _parse_tally(rows, header_i) -> pd.DataFrame:
    records = []
    cur = None
    for row in rows[header_i + 1:]:
        marker = str(row[1]).strip().lower() if len(row) > 1 and row[1] is not None else ""
        if marker in _DRCR:  # a Tally record always starts with a To/By marker
            if cur:
                records.append(cur)
            account = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            vch = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            cur = {"Date": _to_date(row[0]), "Type": str(row[1]).strip(),
                   "Account": account, "VoucherType": vch,
                   "Amount": _largest_number(row), "Details": []}
        else:
            if cur is None:
                continue
            for c in row:
                if c is None:
                    continue
                t = str(c).strip()
                if t and pd.isna(_num(t)):  # text detail only, skip repeated amounts
                    cur["Details"].append(t)
    if cur:
        records.append(cur)

    out = []
    for r in records:
        is_to = r["Type"].lower() == "to"
        out.append({
            "Date": r["Date"],
            "Type": r["Type"],
            "Account": r["Account"],
            "VoucherType": r["VoucherType"],
            "Debit": r["Amount"] if is_to else None,
            "Credit": r["Amount"] if not is_to and r["Type"] else None,
            "Details": "; ".join(dict.fromkeys(r["Details"])),
        })
    return pd.DataFrame(out)


def _parse_clean(rows, header_i) -> pd.DataFrame:
    raw = pd.DataFrame(rows)
    header = [str(c).strip() if c is not None and str(c).strip() else f"col{j}"
              for j, c in enumerate(raw.iloc[header_i])]
    seen = {}
    cols = []
    for c in header:
        if c in seen:
            seen[c] += 1
            c = f"{c}_{seen[c]}"
        else:
            seen[c] = 0
        cols.append(c)
    body = raw.iloc[header_i + 1:].copy()
    body.columns = cols
    body = body.dropna(how="all")
    # Drop embedded column-total rows so sums don't double-count (keep opening/
    # closing balances, which are real entries).
    is_total = body.apply(
        lambda r: any(str(v).strip().lower() in {"total", "grand total"} for v in r),
        axis=1,
    )
    body = body[~is_total].reset_index(drop=True)
    for c in body.columns:
        conv = pd.to_numeric(body[c], errors="coerce")
        if conv.notna().sum() >= body[c].notna().sum() * 0.6:
            body[c] = conv
    return body


def clean_file(path: Path) -> Path:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    out_path = path.with_name(path.stem + "_clean.xlsx")
    with pd.ExcelWriter(str(out_path)) as writer:
        for ws in wb.worksheets:
            rows = _sheet_rows(ws)
            if not rows:
                continue
            h = _find_header(rows)
            meta = _metadata(rows, h)
            if _is_tally(rows, h):
                df = _parse_tally(rows, h)
                kind = "Tally multi-row -> tidied"
            else:
                df = _parse_clean(rows, h)
                kind = "already tabular -> passed through"
            sheet_name = ws.title[:31]
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            deb = pd.to_numeric(df.get("Debit"), errors="coerce").sum() if "Debit" in df else None
            cred = pd.to_numeric(df.get("Credit"), errors="coerce").sum() if "Credit" in df else None
            print(f"  [{ws.title}] {kind}: {len(df)} rows"
                  + (f" | Debit={deb:,.0f} Credit={cred:,.0f}" if deb is not None else "")
                  + (f"\n      meta: {meta}" if meta else ""))
    return out_path


def main(argv):
    targets = argv or ["local_testing/excel"]
    files = []
    for t in targets:
        p = Path(t)
        if p.is_dir():
            files += [f for f in p.iterdir() if f.suffix.lower() in {".xlsx", ".xls"}
                      and "_clean" not in f.stem]
        elif p.suffix.lower() in {".xlsx", ".xls"}:
            files.append(p)
    if not files:
        print("No .xlsx/.xls files found.")
        return
    for f in files:
        print(f"\nCleaning {f} ...")
        out = clean_file(f)
        print(f"  -> wrote {out}")


if __name__ == "__main__":
    main(sys.argv[1:])
